import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def save_to_excel(data, filename="data/history.xlsx", search_date=None):
    """
    Сохраняет данные в Excel файл, создавая один лист на одну дату.
    Применяет стили ко всем ячейкам.
    Статистика добавляется один раз внизу листа и обновляется при каждом добавлении данных.
    Ширина столбцов автоматически подстраивается под содержимое.
    """
    if not data:
        print("Нет данных для сохранения в Excel.")
        return

    df = pd.DataFrame(data)

    # Переименование столбцов в финальные названия
    column_mapping = {
        'time': 'Время отправления рейса с НП',
        'trip_number': '№ рейса',
        'departure_point': 'НП - начальный пункт',
        'arrival_point': 'КП - конечный пункт',
        'carrier': 'Перевозчик',
        'total_seats': 'Всего мест',
        'free_seats': 'Свободно мест',
        'sold_tickets': 'Продано билетов',
        'price': 'Средняя цена за билет на рейсе',
        'source': 'Источник'
    }
    df.rename(columns=column_mapping, inplace=True)

    # Определяем имя листа
    if search_date:
        sheet_name = search_date
    else:
        sheet_name = datetime.now().strftime("%Y-%m-%d")
    sheet_name = sheet_name[:31]  # Ограничение Excel

    # Импортируем стили внутри функции
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    try:
        os.makedirs(os.path.dirname(filename), exist_ok=True)
    except OSError:
        pass

    # Проверяем, существует ли файл и лист
    file_existed = os.path.exists(filename)
    book = None
    if file_existed:
        try:
            book = load_workbook(filename)
            print(f"Открыта существующая книга '{filename}' с листами: {book.sheetnames}")
        except Exception as e:
            print(f"Ошибка при открытии существующего файла {filename}: {e}. Будет создан новый.")
            file_existed = False
            book = None
    # else: book остается None

    try:
        # --- ИСПОЛЬЗУЕМ pd.ExcelWriter с правильными параметрами ---
        # mode='a' для добавления/изменения в существующем файле
        # if_sheet_exists='replace' для замены содержимого конкретного листа
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

            if file_existed and book:
                # Передаем уже загруженную книгу writer'у
                # Это критично для корректной работы с существующими листами
                writer._book = book
                writer._sheets = {}  # Сбрасываем, чтобы избежать конфликтов

            # Проверяем, существует ли лист
            sheet_existed = False
            if book and sheet_name in book.sheetnames:
                sheet_existed = True
                print(f"Обновление существующего листа '{sheet_name}' в книге.")

            if sheet_existed:
                # --- Лист существует, добавляем данные и обновляем статистику ---

                # 1. Считываем уже существующие данные на листе (только сами данные, без заголовков и статистики)
                try:
                    # Читаем данные, начиная со строки 3 (после двух строк заголовков)
                    existing_df = pd.read_excel(filename, sheet_name=sheet_name, header=1, engine='openpyxl')
                    # Удаляем потенциально пустые строки в конце (например, от предыдущей статистики)
                    if not existing_df.empty:
                        # Найдем индекс строки, где в первом столбце появляется "Статистика"
                        stat_indices = existing_df.iloc[:, 0].astype(str).str.startswith("Статистика", na=False)
                        if stat_indices.any():
                            first_stat_index = stat_indices.idxmax()  # Индекс первой строки со статистикой
                            # Удаляем все строки с этой и ниже
                            existing_df = existing_df.iloc[:first_stat_index]
                        # Также удаляем строки, которые полностью пустые (на всякий случай)
                        existing_df.dropna(how='all', inplace=True)

                    print(f"  Прочитано {len(existing_df)} существующих записей со листа '{sheet_name}'.")
                except Exception as e:
                    print(
                        f"  Предупреждение при чтении существующих данных со листа '{sheet_name}': {e}. Считаем, что данных нет.")
                    existing_df = pd.DataFrame()

                # 2. Объединяем старые и новые данные
                df_to_save = pd.concat([existing_df, df], ignore_index=True)
                print(f"  Всего записей для сохранения на лист '{sheet_name}': {len(df_to_save)}.")

                # 3. Записываем данные С ЗАГОЛОВКАМИ на лист (startrow=1)
                # Это заменит содержимое листа, но мы перезапишем его полностью правильно
                df_to_save.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

                # 4. Получаем ссылку на лист из writer'а
                ws = writer.sheets[sheet_name]  # Используем writer.sheets

                # --- Создаем и применяем стили к заголовкам (уже существующая логика) ---
                try:
                    # "Маршрут" объединяет столбцы B-E (2-5)
                    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
                    route_header_cell = ws.cell(row=1, column=2)
                    route_header_cell.value = "Маршрут"
                    route_header_cell.fill = header_fill
                    route_header_cell.font = header_font
                    route_header_cell.alignment = center_alignment

                    # Дата (search_date_str) объединяет столбцы F-H (6-8)
                    if search_date:
                        ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)
                        date_header_cell = ws.cell(row=1, column=6)
                        date_header_cell.value = search_date
                        date_header_cell.fill = header_fill
                        date_header_cell.font = header_font
                        date_header_cell.alignment = center_alignment
                except Exception as e:
                    print(
                        f"Предупреждение при (повторном) создании объединенных заголовков на листе '{sheet_name}': {e}")

                # Применяем стили к основным заголовкам (строка 2)
                for col_num in range(1, len(df_to_save.columns) + 1):
                    cell = ws.cell(row=2, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

                # Применяем стили ко всем ячейкам данных (начиная со строки 3)
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(df_to_save.columns)):
                    for cell in row:
                        cell.border = thin_border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = left_alignment

                # 5. Добавляем обновленную статистику на основе df_to_save
                print("  Добавление обновленной статистики...")
                _add_statistics_to_worksheet(ws, df_to_save, thin_border, center_alignment, left_alignment)

                # 6. Автоматическая ширина столбцов
                _auto_adjust_column_width(ws, df_to_save)

                print(f"Лист '{sheet_name}' успешно ОБНОВЛЕН в книге '{filename}'. Статистика обновлена.")

            else:
                # --- Создаем новый лист в файле (новый или существующий) ---
                print(f"Создание нового листа '{sheet_name}' в книге '{filename}'.")

                # 1. Записываем данные С ЗАГОЛОВКАМИ, начиная со строки 2
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

                # 2. Получаем ссылку на лист
                ws = writer.sheets[sheet_name]

                # --- Создаем и применяем стили к заголовкам ---
                try:
                    # "Маршрут" объединяет столбцы B-E (2-5)
                    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
                    route_header_cell = ws.cell(row=1, column=2)
                    route_header_cell.value = "Маршрут"
                    route_header_cell.fill = header_fill
                    route_header_cell.font = header_font
                    route_header_cell.alignment = center_alignment

                    # Дата (search_date_str) объединяет столбцы F-H (6-8)
                    if search_date:
                        ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)
                        date_header_cell = ws.cell(row=1, column=6)
                        date_header_cell.value = search_date
                        date_header_cell.fill = header_fill
                        date_header_cell.font = header_font
                        date_header_cell.alignment = center_alignment
                except Exception as e:
                    print(f"Предупреждение при создании объединенных заголовков на новом листе '{sheet_name}': {e}")

                # Применяем стили к основным заголовкам (строка 2)
                for col_num in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=2, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border

                # Применяем стили ко всем ячейкам данных (начиная со строки 3)
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = thin_border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = left_alignment

                # 3. Добавляем статистику
                print("  Добавление статистики...")
                _add_statistics_to_worksheet(ws, df, thin_border, center_alignment, left_alignment)

                # 4. Автоматическая ширина столбцов
                _auto_adjust_column_width(ws, df)

                print(f"Новый лист '{sheet_name}' успешно СОЗДАН в книге '{filename}'. Статистика добавлена.")

        # --- После закрытия writer, открываем книгу снова для финальной настройки ---
        # Это необходимо, потому что writer может не применить изменения мгновенно
        try:
            book_final = load_workbook(filename)
            ws_final = book_final[sheet_name]
            # Повторно скорректируем ширину, если нужно
            # _auto_adjust_column_width(ws_final, df if not sheet_existed else df_to_save)
            book_final.save(filename)
            print(f"Финальная настройка книги '{filename}' завершена.")
        except Exception as e:
            print(f"Предупреждение при финальной настройке книги: {e}")

    except Exception as e:
        print(f"Ошибка при сохранении в Excel: {e}")
        import traceback
        traceback.print_exc()

# --- НЕ МЕНЯЕМ: Вспомогательная функция для добавления статистики ---
def _add_statistics_to_worksheet(worksheet, df_for_stats, border_style, center_align, left_align):
    """
    Вспомогательная функция для добавления статистики в лист Excel.
    """
    from openpyxl.styles import Font
    bold_font = Font(bold=True)

    if df_for_stats.empty:
        return

    # Находим строку, с которой начинать статистику (после данных + 2 строки отступа)
    stats_start_row = worksheet.max_row + 3

    # Заголовок статистики
    stat_header_cell = worksheet.cell(row=stats_start_row, column=1, value="Статистика")
    stat_header_cell.font = bold_font
    stat_header_cell.alignment = left_align

    stats_counter = 1
    # 1. Всего рейсов
    total_trips = len(df_for_stats)
    worksheet.cell(row=stats_start_row + stats_counter, column=1, value="Всего рейсов").alignment = left_align
    worksheet.cell(row=stats_start_row + stats_counter, column=2, value=total_trips).alignment = center_align
    stats_counter += 1

    # 2. Средняя цена
    avg_price = df_for_stats['Средняя цена за билет на рейсе'].mean()
    worksheet.cell(row=stats_start_row + stats_counter, column=1, value="Средняя цена (руб.)").alignment = left_align
    worksheet.cell(row=stats_start_row + stats_counter, column=2,
                   value=round(avg_price, 2) if pd.notna(avg_price) else "N/A").alignment = center_align
    stats_counter += 1

    # 3. Направления (Откуда)
    departure_stats = df_for_stats['НП - начальный пункт'].value_counts()
    dep_header_cell = worksheet.cell(row=stats_start_row + stats_counter, column=1, value="Направления (Откуда)")
    dep_header_cell.font = bold_font
    dep_header_cell.alignment = left_align
    stats_counter += 1
    for dep_point, count in departure_stats.items():
        worksheet.cell(row=stats_start_row + stats_counter, column=1, value=f"  {dep_point}").alignment = left_align
        worksheet.cell(row=stats_start_row + stats_counter, column=2, value=count).alignment = center_align
        stats_counter += 1

    # 4. Прибытие (Куда)
    arrival_stats = df_for_stats['КП - конечный пункт'].value_counts()
    arr_header_cell = worksheet.cell(row=stats_start_row + stats_counter, column=1, value="Прибытие (Куда)")
    arr_header_cell.font = bold_font
    arr_header_cell.alignment = left_align
    stats_counter += 1
    for arr_point, count in arrival_stats.items():
        worksheet.cell(row=stats_start_row + stats_counter, column=1, value=f"  {arr_point}").alignment = left_align
        worksheet.cell(row=stats_start_row + stats_counter, column=2, value=count).alignment = center_align
        stats_counter += 1

    # 5. Перевозчики
    carrier_stats = df_for_stats['Перевозчик'].value_counts()
    carr_header_cell = worksheet.cell(row=stats_start_row + stats_counter, column=1, value="Перевозчики")
    carr_header_cell.font = bold_font
    carr_header_cell.alignment = left_align
    stats_counter += 1
    for carrier, count in carrier_stats.items():
        if carrier != "Не определен":
            worksheet.cell(row=stats_start_row + stats_counter, column=1, value=f"  {carrier}").alignment = left_align
            worksheet.cell(row=stats_start_row + stats_counter, column=2, value=count).alignment = center_align
            stats_counter += 1

    # Применяем границы и выравнивание к ячейкам статистики
    for row in worksheet.iter_rows(min_row=stats_start_row, max_row=stats_start_row + stats_counter - 1,
                                   min_col=1, max_col=2):
        for cell in row:
            cell.border = border_style

# --- НЕ МЕНЯЕМ: Функция автоматической ширины столбцов ---
def _auto_adjust_column_width(worksheet, dataframe):
    """
    Автоматически подстраивает ширину столбцов под содержимое заголовков и данных.
    """
    for col_idx, col_name in enumerate(dataframe.columns, 1):
        # Ширина по заголовку
        column_letter = worksheet.cell(row=2, column=col_idx).column_letter # Основные заголовки во 2й строке
        max_length = len(str(col_name))

        # Проверяем ширину по первым 100 строкам данных для производительности
        for row_idx in range(3, min(103, worksheet.max_row + 1)): # Начинаем с 3й строки данных
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                try:
                    cell_length = len(str(cell_value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass # Игнорируем ошибки

        # Устанавливаем ширину, добавляя небольшой отступ
        adjusted_width = min(max_length + 2, 50) # Ограничиваем максимальную ширину
        worksheet.column_dimensions[column_letter].width = adjusted_width
    print("  Ширина столбцов автоматически скорректирована.")
